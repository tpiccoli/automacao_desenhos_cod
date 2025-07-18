# -*- coding: utf-8 -*-
"""
Created on Thu Jan 30 12:05:16 2025

@author: tiago.piccoli

#versao A
"""

import pandas as pd
import os
import shutil
from openpyxl import load_workbook
#import time
#from datetime import date
#from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
global atencao 
atencao=[]

#funções------------------------------------------------------------------------
def mover_arquivos(subst_des, pasta_fabrica, pasta_obsoletos, pasta_novos): #desenhos existentes    
    for arquivo in subst_des:
        # Procurar arquivos com o mesmo nome (considerar quaisquer extensões)
        #arquivo_str = str(arquivo)
        for extensao in ['.dxf', '.PDF', '.x_t']:
            arquivo_fabrica = os.path.join(pasta_fabrica, arquivo + extensao)
            arquivo_novos = os.path.join(pasta_novos, arquivo + extensao)
            arquivo_obsoletos = os.path.join(pasta_obsoletos, arquivo + extensao)
 
            # Mover da pasta desenhos fabrica para pasta obsoletos
            if os.path.exists(arquivo_fabrica):
                if extensao == '.dxf' and not os.path.exists(os.path.join(pasta_novos, arquivo + '.dxf')):
                    # Se não houver novo arquivo .dxf, não mover para obsoletos
                    print(f"Arquivo {arquivo}.dxf permanece na pasta Desenhos Fábrica")
                elif extensao == '.x_t' and not os.path.exists(os.path.join(pasta_novos, arquivo + '.x_t')):
                    # Se não houver novo arquivo .x_t, não mover para obsoletos
                    print(f"Arquivo {arquivo}.x_t permanece na pasta Desenhos Fábrica")
                else:
                   try: 
                       shutil.move(arquivo_fabrica, arquivo_obsoletos)
                   except:
                       print(arquivo, "Não pode ser substituído")
                       atencao.append(arquivo)
                       continue
                   
            # Copiar da pasta novos desenhos para pasta desenhos fabrica
            if os.path.exists(arquivo_novos):
                try:
                    shutil.copy2(arquivo_novos, arquivo_fabrica)
                    print(arquivo," copiado para Desenhos Fabrica")
                    os.remove(arquivo_novos)# Deletar da pasta novos desenhos
                except:
                    print(arquivo," Não pode ser substituído")
                    atencao.append(arquivo)
                    continue
                
            elif extensao == ".x_t": #verificacao de sufixo _0
                arquivo_novos_0 = os.path.join(pasta_novos, arquivo + "_0" + extensao)
                if os.path.exists(arquivo_novos_0):
                    try:
                        shutil.copy2(arquivo_novos_0, arquivo_fabrica)
                        if os.path.exists(arquivo_novos):
                            os.remove(arquivo_novos)
                        os.rename(arquivo_novos_0, arquivo_fabrica)
                        print(arquivo, 'Movido para pasta Desenhos Fábrica')      
                    except:
                        print(arquivo, "Não pode ser substituído")
                        atencao.append(arquivo)
                        continue
                else:
                    pass
            else:
                pass

def novo_arquivos(mover_des, pasta_fabrica, pasta_novos): #desenhos novos
    for arquivo in mover_des: #para cada arquivo da lista de desenhos novos
        # Procurar arquivos com o mesmo nome (considerar quaisquer extensões)
        arquivo_str = str(arquivo)
        for extensao in ['.dxf', '.PDF', '.x_t']:
            arquivo_fabrica = os.path.join(pasta_fabrica, arquivo_str + extensao)
            arquivo_novos = os.path.join(pasta_novos, arquivo_str + extensao)
            
            # Copiar da pasta novos desenhos para pasta desenhos fabrica
            if os.path.exists(arquivo_novos):
                try:
                    shutil.copy2(arquivo_novos, arquivo_fabrica)
                    print(arquivo, 'Movido para pasta Desenhos Fábrica')
                    os.remove(arquivo_novos)
                except:
                    print(arquivo," Não pode ser substituído")
                    atencao.append(arquivo)
                    continue
            elif extensao == ".x_t": #verificacao de sufixo _0
                arquivo_novos_0 = os.path.join(pasta_novos, arquivo_str + "_0" + extensao)
                if os.path.exists(arquivo_novos_0):
                    try:
                        shutil.copy2(arquivo_novos_0, arquivo_fabrica)
                        if os.path.exists(arquivo_novos):
                            os.remove(arquivo_novos)
                        os.rename(arquivo_novos_0, arquivo_fabrica)
                        print(arquivo, 'Movido para pasta Desenhos Fábrica')
                    except:
                        print(arquivo, "Não pode ser substituído")
                        atencao.append(arquivo)
                        continue
                else:
                    pass
            else:
                pass

def ler_planilha(caminho_planilha):
    while True:
        try:
            df = pd.read_excel(caminho_planilha)
            print("Planilha lida com sucesso!")
            return df
        except PermissionError:
            input("Por favor, abra a planilha, preencha a coluna PROCESSOS a partir do roteiro do item, salve, feche e tecle Enter quando tiver concluído.")

# Função para salvar lista em arquivo .txt
def emails_lista(lista, nome_arquivo):
    with open(f'{nome_arquivo}.txt', 'w') as f:
        for item in lista:
            f.write(f"{item}\n")

def listar_des_nv(pasta_novos):
    arquivos_novos = []
    for arquivo in os.listdir(pasta_novos):
        if arquivo.lower().endswith('.pdf'):
            arquivos_novos.append(os.path.splitext(arquivo)[0])
            print('desenho pode ser novo: ', arquivo)
    return arquivos_novos

def des_nv_ant(arquivos_pdf, pasta_fabrica): #verifica se o desenho novo já existe na pasta.
    arquivos_existentes = []
    arquivos_a_remover = []
    for arquivo in arquivos_novos:
        caminho_arquivo = os.path.join(pasta_fabrica, arquivo+".pdf")
        caminho_arquivo2 = os.path.join(pasta_fabrica, arquivo+".PDF")
        
        if os.path.exists(caminho_arquivo) or os.path.exists(caminho_arquivo2):
            arquivos_existentes.append(arquivo)
            arquivos_a_remover.append(arquivo)
            print('desenho não é novo, vai ser movido: ',arquivo)
        else:
            print('desenho é realmente novo: ',arquivo)
            pass
    for arquivo in arquivos_a_remover:
        arquivos_novos.remove(arquivo)
        
    return arquivos_existentes
#---------------------------------------------------

print("PROGRAMA FLUXO PARA LIBERAÇÃO DE DESENHOS, Versão A")
input("Tecle 'Enter' para iniciar o programa")

pasta_fabrica = 'P:\\Útil\\Desenhos Tecnicos'
pasta_obsoletos = 'P:\\Útil\\Desenhos Tecnicos\\OBSOLETOS'
pasta_novos = 'P:\\Útil\\Desenhos Tecnicos\\Novos Desenhos'

# pasta_fabrica = r'C:\Users\tiago.piccoli\des_fabrica'
# pasta_obsoletos = r'C:\Users\tiago.piccoli\des_obsletos'
# pasta_novos = r'C:\Users\tiago.piccoli\des_nv'

#dia=date.today()
#data_hj=dia.strftime('%d/%m/%Y')
arquivos_novos = listar_des_nv(pasta_novos) #gera lista com os arquivos novos
arquivos_existentes = des_nv_ant(arquivos_novos, pasta_fabrica) #verifica se é desenho novo

# Para os arquivos revisados (REVISADO)
novos_dados_rev = pd.DataFrame([{
    'CODIGO': arquivo, 
    #'Data Liberação': data_hj,
    'PROCESSOS':'',    
    'VERSAO': 'REVISADO',
    'FLUXO': 0,
    'OBS':''
} for arquivo in arquivos_existentes])

# Para os arquivos novos (DESENHO NOVO)
novos_dados_nv = pd.DataFrame([{
    #'CODIGO': os.path.splitext(arquivo)[0],
    #'Data Liberação': data_hj,
    'CODIGO': arquivo,
    'PROCESSOS':'',    
    'VERSAO': 'DESENHO NOVO',
    'FLUXO': 0,
    'OBS':''
} for arquivo in arquivos_novos])

# Concatenar o DF dos arquivos novos e dos revisados
df = pd.concat([novos_dados_nv, novos_dados_rev], ignore_index=True)

# Salvar o DataFrame no Excel
df.to_excel('FLUXO DESENHOS.xlsx', index=False, engine='openpyxl')

# Carregar o workbook e a planilha
wb = load_workbook('FLUXO DESENHOS.xlsx')
ws = wb.active

# Definir as opções da lista suspensa
opcoes_lista = [
    "LASER", "LASER e DOBRA", "USINAGEM", "SERRA", "MONTAGEM", "PRE-MONTAGEM",
    "TERCEIROS", "ALIMENTADORES", "SOLDA", "FONTES", "COMPRADO", "CENTRO",
    "OUTROS", "SERRA e CENTRO", "ROSQUEAMENTO", "GRAVADOR", "SERRA e TORNO",
    "SERRA - TORNO - CENTRO", "DOBRA", "SEM ROTEIRO"
]

# Criar a validação de dados (drop-down)
dv = DataValidation(
    type="list",
    formula1=f'"{",".join(opcoes_lista)}"',
    allow_blank=True,
    showDropDown=True
)

# Validação de dados à coluna "PROCESSOS"
coluna_processos = ws['B']
for cell in coluna_processos[1:]:
    dv.add(cell)

# Adicionar a validação à worksheet
ws.add_data_validation(dv)

# Salvar o workbook atualizado
wb.save('FLUXO DESENHOS.xlsx')
print('-------------------------------------------------')
print('Relatório para preenchimento de PROCESSO/OPERAÇÃO')

input('Por favor, abra a planilha, preencha a coluna "PROCESSOS" a partir do roteiro do item, salve, feche e tecle Enter quando tiver concluído.')

caminho_planilha = "FLUXO DESENHOS.XLSX"
df = ler_planilha(caminho_planilha)

cod_qualidade = []
cod_compras = []
cod_laser = []
ver_impressao_cq = []
atencao=[]
sem_roteiro=[]

mover_des=[] #lista para função de operação de desenhos novos
subst_des=[] #lista para função de operação de desenhos antigos

df['OBS'] = df['OBS'].astype('object')

for index, row in df.iterrows(): #Algoritmo
    fluxo = row['FLUXO']
    codigo = str(row['CODIGO'])
    versao = row['VERSAO']
    processos = row['PROCESSOS']
    obs=row['OBS']
    
    #verificar se é desenho novo, vendo se há existência do arquivo na pasta.
    
    if codigo in arquivos_novos:
        
        if processos in ('COMPRADO','TERCEIROS'):
            cod_compras.append(codigo)
            df.at[index, 'FLUXO'] = 1
            df.at[index, 'OBS']="Desenho Movido"
            #df.at[index, 'Data Liberação']=data_hj
            mover_des.append(codigo)
            continue
            
        elif processos in ('MONTAGEM', 'PRE-MONTAGEM', 'ALIMENTADORES', 'SERRA','FONTES','GRAVADOR'):
            df.at[index, 'FLUXO'] = 1
            df.at[index, 'OBS']="Desenho Movido"
            #df.at[index, 'Data Liberação']=data_hj
            mover_des.append(codigo)
            continue
        
        elif processos in ('LASER','LASER e DOBRA'):
            cod_laser.append(codigo)
            ver_impressao_cq.append(codigo)
            df.at[index, 'FLUXO'] = 1
            df.at[index, 'OBS']="Desenho Movido, Verificar CQ e Imprimir"
            #df.at[index, 'Data Liberação']=data_hj
            mover_des.append(codigo)
            continue
        
        elif processos in ('USINAGEM','SOLDA','SERRA e CENTRO','SERRA - TORNO - CENTRO','SERRA e TORNO'):
            ver_impressao_cq.append(codigo)
            df.at[index, 'FLUXO'] = 1      
            df.at[index, 'OBS']="Desenho Movido, Verificar CQ e Imprimir"
            #df.at[index, 'Data Liberação']=data_hj
            mover_des.append(codigo)
            continue

        elif processos in ('SEM ROTEIRO'):
            sem_roteiro.append(codigo)
            df.at[index, 'FLUXO'] = 1      
            df.at[index, 'OBS']="Desenho Movido, definir operação e ações do item"
            #df.at[index, 'Data Liberação']=data_hj
            mover_des.append(codigo)
            continue
        else:
            atencao.append(codigo)
            df.at[index, 'OBS']="Desenho NÃO MANIPULADO por não se enquadrar em nenhum processo, VERIFICAR"
            continue
    elif codigo in arquivos_existentes:
        
        if processos in ("LASER","LASER e DOBRA"):
            cod_laser.append(codigo)
            ver_impressao_cq.append(codigo)
            df.at[index, 'FLUXO'] = 1
            df.at[index, 'OBS']="Desenho Substituído, Verificar/Atualizar CQ caso haja e Imprimir"
            #df.at[index, 'Data Liberação']=data_hj
            subst_des.append(codigo)
            continue
            
        elif processos in ("TERCEIROS","COMPRADO"):
            cod_qualidade.append(codigo)
            df.at[index, 'FLUXO'] = 1
            df.at[index, 'OBS']="Desenho Substituído, E-mail Suprimentos e Qualidade"
            #df.at[index, 'Data Liberação']=data_hj
            subst_des.append(codigo)
            continue
        
        elif processos in ('USINAGEM','SOLDA','SERRA e CENTRO','SERRA - TORNO - CENTRO','SERRA e TORNO'):
            ver_impressao_cq.append(codigo)
            df.at[index, 'FLUXO'] = 1
            df.at[index, 'OBS']="Desenho Substituído, Verificar/Atualizar CQ caso haja e Imprimir"
            #df.at[index, 'Data Liberação']=data_hj
            subst_des.append(codigo)
            continue
        
        elif processos in ('MONTAGEM', 'PRE-MONTAGEM', 'ALIMENTADORES', 'SERRA','FONTES','GRAVADOR'):
            df.at[index, 'FLUXO'] = 1
            df.at[index, 'OBS']="Desenho Substituído"
            #df.at[index, 'Data Liberação']=data_hj
            subst_des.append(codigo)
            continue

        elif processos in ('SEM ROTEIRO'):
            sem_roteiro.append(codigo)
            df.at[index, 'FLUXO'] = 1      
            df.at[index, 'OBS']="Desenho Substituído, definir operação e ações do item"
            #df.at[index, 'Data Liberação']=data_hj
            subst_des.append(codigo)
            continue

        else:
            atencao.append(codigo)
            df.at[index, 'OBS']="Desenho NÃO MANIPULADO por não se enquadrar em nenhum processo, VERIFICAR"
            continue
    else:
        atencao.append(codigo)
        df.at[index, 'OBS']="Desenho ignorado pelo algoritmo, VERIFICAR"
        continue

print('------------------------------------------------')
novo_arquivos(mover_des, pasta_fabrica, pasta_novos)
mover_arquivos(subst_des, pasta_fabrica, pasta_obsoletos, pasta_novos)

df.to_excel('FLUXO DESENHOS.xlsx',index=False, engine='openpyxl')
print('--------------RESUMO DAS OPERAÇÕES--------------')
print(df)

emails_lista(cod_compras, 'cod_compras')
emails_lista(cod_laser, 'cod_laser')
emails_lista(cod_qualidade, 'cod_qualidade')
emails_lista(ver_impressao_cq, 'ver_impressao_cq')
emails_lista(atencao, 'VERIFICAR')
emails_lista(sem_roteiro, 'sem_roteiro')

input()
